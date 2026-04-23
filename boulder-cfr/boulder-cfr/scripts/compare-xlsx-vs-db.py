"""Compare CFR Detail rows in xlsx vs parsed_debits.json + parsed_credits.json
(what DB was seeded with). Reports:
 - counts
 - per-division mismatch
 - rows in xlsx but not in parsed
 - rows in parsed but not in xlsx
 - field-level diffs (dates, amounts, vendor, bid item, etc.)
"""
import openpyxl, json
from pathlib import Path
from collections import defaultdict

XLSX = Path(r"C:\Users\Offic\Dropbox\CFR-DRAW\CFR v31.xlsx")
PARSED_DIR = Path(__file__).parent / "extracted"
DEBITS = PARSED_DIR / "parsed_debits.json"
CREDITS = PARSED_DIR / "parsed_credits.json"

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["Detail"]

# Columns (1-based): 3=Date, 4=G703, 5=Draw, 6=Description, 7=Commentary,
# 8=DebitG, 9=DebitR, 10=DebitN, 11=CreditG, 12=CreditR, 13=CreditN,
# 14=BidItem, 15=Counterparty, 16=PaidBy, 17=Backup, 18=ReceivedK1, 19=Type

def cents(v):
    if v is None or v == "": return 0
    try: return round(float(v) * 100)
    except: return 0

def iso_date(v):
    if v is None: return None
    if hasattr(v, "date"): return v.date().isoformat()
    return str(v)

xlsx_debits = []
xlsx_credits = []
current_div = None

for i in range(6, ws.max_row + 1):
    row = [ws.cell(row=i, column=c).value for c in range(1, 22)]
    date_cell = row[2]
    if isinstance(date_cell, str) and date_cell and date_cell[0].isdigit() and ". " in date_cell:
        try: current_div = int(date_cell.split(".")[0])
        except: pass
        continue
    if current_div is None: continue
    desc = row[5]
    if not desc: continue

    dg = cents(row[7])
    dn = cents(row[9])
    cg = cents(row[10])
    cn = cents(row[12])

    rec = {
        "divNum": current_div,
        "drawNum": int(row[4]) if isinstance(row[4], (int, float)) else None,
        "date": iso_date(date_cell),
        "description": (desc or "").strip(),
        "bidItem": (row[13] or "").strip(),
        "counterparty": (row[14] or "").strip(),
    }
    if dg or dn:
        xlsx_debits.append({**rec, "grossCents": dg, "netCents": dn})
    if cg or cn:
        xlsx_credits.append({**rec, "grossCents": cg, "netCents": cn})

parsed_debits = json.loads(DEBITS.read_text())
parsed_credits = json.loads(CREDITS.read_text())

def make_key(r):
    return (r["divNum"], r["drawNum"], r["date"] or "", (r["description"] or "").strip(), r["grossCents"])

x_dset = {make_key(r) for r in xlsx_debits}
p_dset = {make_key(r) for r in parsed_debits}
x_cset = {make_key(r) for r in xlsx_credits}
p_cset = {make_key(r) for r in parsed_credits}

print("="*70)
print("DEBITS")
print("="*70)
print(f"xlsx:   {len(xlsx_debits)}")
print(f"parsed: {len(parsed_debits)}")
print(f"match:  {len(x_dset & p_dset)}")
only_x = x_dset - p_dset
only_p = p_dset - x_dset
print(f"only in xlsx:   {len(only_x)}")
print(f"only in parsed: {len(only_p)}")

if only_x:
    print("\nFirst 10 rows in xlsx but NOT in parsed:")
    for k in list(only_x)[:10]:
        print(f"  div={k[0]} draw={k[1]} date={k[2]} desc={k[3][:40]!r} gross={k[4]/100:.2f}")

if only_p:
    print("\nFirst 10 rows in parsed but NOT in xlsx:")
    for k in list(only_p)[:10]:
        print(f"  div={k[0]} draw={k[1]} date={k[2]} desc={k[3][:40]!r} gross={k[4]/100:.2f}")

# Per-division summary
print("\nPer-division debit totals:")
print(f"{'div':>4} {'xlsx_n':>8} {'parsed_n':>9} {'xlsx_sum':>14} {'parsed_sum':>14} {'delta':>12}")
x_by_div = defaultdict(lambda: [0, 0])
p_by_div = defaultdict(lambda: [0, 0])
for r in xlsx_debits:
    x_by_div[r["divNum"]][0] += 1
    x_by_div[r["divNum"]][1] += r["grossCents"]
for r in parsed_debits:
    p_by_div[r["divNum"]][0] += 1
    p_by_div[r["divNum"]][1] += r["grossCents"]
for div in sorted(set(list(x_by_div.keys()) + list(p_by_div.keys()))):
    xn, xs = x_by_div.get(div, [0, 0])
    pn, ps = p_by_div.get(div, [0, 0])
    mark = "" if (xn == pn and xs == ps) else "  ← MISMATCH"
    print(f"{div:>4} {xn:>8} {pn:>9} {xs/100:>14.2f} {ps/100:>14.2f} {(ps-xs)/100:>+12.2f}{mark}")

print("\n" + "="*70)
print("CREDITS")
print("="*70)
print(f"xlsx:   {len(xlsx_credits)}")
print(f"parsed: {len(parsed_credits)}")
print(f"match:  {len(x_cset & p_cset)}")
print(f"only in xlsx:   {len(x_cset - p_cset)}")
print(f"only in parsed: {len(p_cset - x_cset)}")

# Grand totals
x_total = sum(r["grossCents"] for r in xlsx_debits)
p_total = sum(r["grossCents"] for r in parsed_debits)
print("\n" + "="*70)
print("GRAND TOTALS")
print("="*70)
print(f"Debits   xlsx: ${x_total/100:>15,.2f}   parsed: ${p_total/100:>15,.2f}   delta ${(p_total-x_total)/100:>+12,.2f}")
x_total_c = sum(r["grossCents"] for r in xlsx_credits)
p_total_c = sum(r["grossCents"] for r in parsed_credits)
print(f"Credits  xlsx: ${x_total_c/100:>15,.2f}   parsed: ${p_total_c/100:>15,.2f}   delta ${(p_total_c-x_total_c)/100:>+12,.2f}")
