"""Rebuild parsed_debits.json + parsed_credits.json from CFR v31.xlsx Detail sheet.
Includes receivedK1 field (missing in prior parse)."""
import openpyxl, json, sys
from pathlib import Path

XLSX = Path(r"C:\Users\Offic\Dropbox\CFR-DRAW\CFR v31.xlsx")
OUT_DEBITS = Path(__file__).parent / "extracted" / "parsed_debits.json"
OUT_CREDITS = Path(__file__).parent / "extracted" / "parsed_credits.json"

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb["Detail"]

# Column indexes (1-based): 3=Date, 4=G703, 5=Draw, 6=Description, 7=Commentary,
# 8=DebitG, 9=DebitR, 10=DebitN, 11=CreditG, 12=CreditR, 13=CreditN,
# 14=BidItem, 15=Counterparty, 16=PaidBy, 17=Backup, 18=ReceivedK1, 19=Type
COL = {"date": 3, "g703": 4, "draw": 5, "desc": 6, "comm": 7,
       "dg": 8, "dr": 9, "dn": 10, "cg": 11, "cr": 12, "cn": 13,
       "bidItem": 14, "counterparty": 15, "paidBy": 16, "backup": 17,
       "receivedK1": 18, "type": 19}

def val(row, key): return row[COL[key] - 1]

def cents(v):
    if v is None or v == "": return 0
    try: return round(float(v) * 100)
    except: return 0

def iso_date(v):
    if v is None: return None
    if hasattr(v, "date"): return v.date().isoformat()
    return str(v)

debits = []
credits = []
current_div = None

for i in range(6, ws.max_row + 1):
    row = [ws.cell(row=i, column=c).value for c in range(1, 22)]
    # Division header row: col 3 (Date) holds "N. Name"
    date_cell = val(row, "date")
    if isinstance(date_cell, str) and date_cell and date_cell[0].isdigit() and ". " in date_cell:
        try:
            current_div = int(date_cell.split(".")[0])
        except:
            pass
        continue
    if current_div is None: continue
    desc = val(row, "desc")
    if not desc: continue

    dg, dr, dn = cents(val(row, "dg")), cents(val(row, "dr")), cents(val(row, "dn"))
    cg, cr, cn = cents(val(row, "cg")), cents(val(row, "cr")), cents(val(row, "cn"))

    common = {
        "divNum": current_div,
        "drawNum": int(val(row, "draw")) if isinstance(val(row, "draw"), (int, float)) else None,
        "g703": int(val(row, "g703")) if isinstance(val(row, "g703"), (int, float)) else None,
        "date": iso_date(date_cell),
        "description": desc or "",
        "commentary": val(row, "comm") or "",
        "bidItem": val(row, "bidItem") or "",
        "counterparty": val(row, "counterparty") or "",
        "paidBy": val(row, "paidBy") or "",
        "backup": str(val(row, "backup")) if val(row, "backup") is not None else "",
        "receivedK1": val(row, "receivedK1") or "",
        "type": val(row, "type") or "",
    }

    if dg or dn:
        debits.append({**common, "grossCents": dg, "retainageCents": dr, "netCents": dn})
    if cg or cn:
        credits.append({**common, "grossCents": cg, "retainageCents": cr, "netCents": cn})

OUT_DEBITS.write_text(json.dumps(debits, indent=2))
OUT_CREDITS.write_text(json.dumps(credits, indent=2))
print(f"debits: {len(debits)}")
print(f"credits: {len(credits)}")
rk = sum(1 for d in debits if d["receivedK1"])
print(f"debits with receivedK1: {rk}")
bi = sum(1 for d in debits if d["bidItem"])
print(f"debits with bidItem: {bi}")
